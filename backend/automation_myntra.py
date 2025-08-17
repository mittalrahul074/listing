from selenium.webdriver.common.by import By
import time, os
import random
import pyautogui
import undetected_chromedriver as uc
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, ElementClickInterceptedException
import openpyxl
from difflib import get_close_matches
import shutil
from sku_tracker import mark_completed, is_completed, all_completed

PROFILE_DIR = os.path.join(os.path.dirname(__file__), "myntra_profile")

def get_driver():
    chrome_options = uc.ChromeOptions()
    chrome_options.add_argument(f"--user-data-dir={PROFILE_DIR}")
    driver = uc.Chrome(options=chrome_options)
    return driver
    
def set_up_excel(product_data):
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    excel_path = os.path.join(base_dir, 'myntra_excel', f'{product_data["productType"]}.xlsx')

    output_filename = f"{product_data['sku']}.xlsx"
    output_path = os.path.join(base_dir, 'myntra_sku', output_filename)
    shutil.copyfile(excel_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    sheet = wb['Necklace and Chains']
    header_row = sheet[3]
    column_map = {}
    for idx, cell in enumerate(header_row, 1):
        if cell.value:
            column_map[cell.value.strip()] = idx
    row =4
    sheet.cell(row, column=1, value=1)    
    sheet.cell(row, column=2, value=1)    
    sheet.cell(row, column=3, value=f"{product_data['sku']}")    
    sheet.cell(row, column=4, value=f"{product_data['productName']}")    
    sheet.cell(row, column=5, value=f"{product_data['productName']}")
    sheet.cell(row, column=6, value=f"Gorkhastyle")
    sheet.cell(row, column=7, value="Mittal Distributors, 1st floor niladri galaxy, bidhan market, darjeeling, 734001")
    sheet.cell(row, column=8, value="Mittal Distributors, 1st floor niladri galaxy, bidhan market, darjeeling, 734001")
    sheet.cell(row, column=10, value="India")
    sheet.cell(row, column=15, value="Necklace and Chains")
    sheet.cell(row, column=16, value="28 Inches")
    sheet.cell(row, column=17, value="Onesize")
    sheet.cell(row, column=18, value="Yes")
    sheet.cell(row, column=19, value=f"{product_data['color']}")
    sheet.cell(row, column=21, value="71171990")
    sheet.cell(row, column=23, value=f"{product_data['product_mrp']}")
    sheet.cell(row, column=24, value=f"Adults-Women")
    sheet.cell(row, column=25, value=f"{product_data['color']}")
    sheet.cell(row, column=28, value=f"Fashion")
    sheet.cell(row, column=29, value=f"{product_data['occasion']}")
    #enter current year in YYYY format for column 30
    current_year = time.strftime("%Y")
    sheet.cell(row, column=30, value=current_year)
    #enter season value i.e. Spring,Summer, Fall or Winter as per the current month in column 31
    month = time.strftime("%m")
    if month in ['03', '04', '05']:
        season = "Spring"
    elif month in ['06', '07', '08']:
        season = "Summer"
    elif month in ['09', '10', '11']:
        season = "Fall"
    else:
        season = "Winter"
    sheet.cell(row, column=31, value=season)
    sheet.cell(row, column=32, value=f"{product_data['description']}")
    sheet.cell(row, column=34, value=f"Material: Alloy Care Instructions: Wipe your jewellery with a soft cloth after every use Always store your jewellery in a flat box to avoid accidental scratches Keep sprays and perfumes away from your jewellery Do not soak your jewellery in soap water Clean your jewellery using a soft brush Dipped in jewellery cleaning solution only")
    sheet.cell(row, column=35, value=f"Lenght - 28 inches")
    sheet.cell(row, column=36, value=f"{product_data['productName']}")
    sheet.cell(row, column=41, value=f"{product_data['material']}")
    sheet.cell(row, column=42, value=f"{product_data['stoneType']}")
    sheet.cell(row, column=45, value="NA")
    sheet.cell(row, column=47, value="6 months")
    sheet.cell(row, column=48, value=f"{product_data['plating']}")
    sheet.cell(row, column=49, value=f"Handcrafted")
    sheet.cell(row, column=50, value=f"Regular")
    sheet.cell(row, column=51, value=f"Pieces")
    sheet.cell(row, column=53, value=f"mittal.distributor.online@gmail.com  or +919382629271")
    sheet.cell(row, column=57, value=f"{product_data['inventory']}")

    wb.save(output_path)

    print(f"✅ Product added to '{output_path}' in sheet 'Necklace and Chains'")
    return output_path

def open_myntra_website_and_upload(excel_path, sku_folder, sku):
    driver = None
    try:
        driver = get_driver()
        wait = WebDriverWait(driver, 15)
        driver.get("https://partners.myntrainfo.com/DiyCataloguingV2")
        time.sleep(15)  # Wait for the page to load

        if "login" in driver.current_url.lower():
            input("Log in on Myntra, then press Enter...")
            while "login" in driver.current_url.lower():
                time.sleep(5)
        
        driver.get("https://partners.myntrainfo.com/DiyCataloguingV2")
        time.sleep(5)  # Wait for the page to load

        upload_button = driver.find_element(By.XPATH, '//button[normalize-space()="Add New DIY Products"]')
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", upload_button)
        upload_button.click()
        time.sleep(3)



        
    except TimeoutException:
        print("Login page did not load in time.")
        driver.quit()
        return
    
def automate_myntra_listing(product_data, sku_folder):
    excel_path = set_up_excel(product_data)
    open_myntra_website_and_upload(excel_path, sku_folder, product_data['sku'])
    mark_completed(product_data['sku'], "myntra")
    return
    