"""
Myntra Excel Export Module
==========================

Handles automatic Excel file generation for Myntra bulk product upload.
Includes photo upload reminder and workflow automation.
"""

import pandas as pd
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime, timedelta
import time
import threading
import json
from dataclasses import asdict

from data_models import Product, ProductVariant, Material, ProductType
from image_manager import ImageManager


# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Myntra Excel template columns (based on jewelry category)
MYNTRA_TEMPLATE_COLUMNS = [
    # Basic Information
    'Style ID',           # SKU
    'Product Name',       # Product title
    'Brand Name',         # Brand
    'Article Type',       # Necklace, Bangle, etc.
    'Product Details',    # Description

    # Pricing
    'MRP',               # Maximum Retail Price
    'Best Price',        # Selling price

    # Physical Properties
    'Primary Material',   # Material
    'Brand Color',       # Color
    'Plating',          # Gold Plated, Silver Plated
    'Base Metal',       # Brass, Alloy, etc.
    'Stone Type',       # If applicable
    'Net Weight',       # In grams
    'Gross Weight',     # In grams

    # Size and Dimensions
    'Size',             # Size specification
    'Length',           # For necklaces
    'Width',            # Width measurement
    'Height',           # Height measurement

    # Classification
    'Gender',           # Men/Women/Unisex
    'Category',         # Jewelry
    'Usage',            # Casual, Party, Wedding
    'Collection',       # Collection name
    'Season',           # All Season, Winter, Summer
    'Year',             # Manufacturing year

    # Business Details
    'HSN',              # HSN Code
    'Net Quantity',     # Quantity
    'Inventory',        # Stock quantity

    # Marketing
    'Search Tags',      # Comma separated tags
    'Occasion',         # Wedding, Party, Casual
    'Trend',           # Traditional, Modern, etc.

    # Origin and Manufacturing
    'Country of Origin', # India, China, etc.
    'Manufacturer',     # Manufacturer name
    'Packer',          # Packer name
    'Importer',        # If imported

    # Additional
    'Warranty',         # Warranty period in months
    'Certification',    # BIS, Hallmark, etc.
    'Care Instructions', # Care and maintenance

    # Image URLs (to be filled after upload)
    'Image 1',
    'Image 2',
    'Image 3',
    'Image 4',
    'Image 5',
    'Image 6',
    'Image 7',
    'Image 8',
    'Image 9',
    'Image 10'
]


class MyntraExporter:
    """Handles Excel generation and upload workflow for Myntra."""

    def __init__(self, output_dir: Path = None):
        """Initialize the Myntra exporter."""
        self.output_dir = Path(output_dir) if output_dir else Path("exports/myntra")
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.image_manager = ImageManager()
        self.upload_reminders = {}  # Track photo upload reminders

        logger.info(f"MyntraExporter initialized - Output: {self.output_dir}")

    def product_to_myntra_row(self, product: Product, variant: ProductVariant = None) -> Dict[str, Any]:
        """
        Convert a Product (and optional variant) to Myntra Excel row data.

        Args:
            product: Product object
            variant: Optional variant for bangle sizes etc.

        Returns:
            Dictionary mapping column names to values
        """
        # Use variant SKU if provided, otherwise main product SKU
        sku = variant.sku if variant else product.sku

        # Use variant inventory if provided
        inventory = variant.inventory if variant else product.inventory
        size = variant.size if variant else product.size

        # Convert material enum to string
        material_str = product.material.value.replace('_', ' ').title() if isinstance(product.material, Material) else str(product.material)

        # Format search tags
        search_tags = ', '.join(product.search_tags) if product.search_tags else ''

        row_data = {
            # Basic Information
            'Style ID': sku,
            'Product Name': product.product_name,
            'Brand Name': product.brand,
            'Article Type': self._get_myntra_article_type(product.product_type),
            'Product Details': product.description,

            # Pricing
            'MRP': product.mrp,
            'Best Price': product.selling_price,

            # Physical Properties
            'Primary Material': material_str,
            'Brand Color': product.color,
            'Plating': product.plating,
            'Base Metal': product.base_metal,
            'Stone Type': product.stone_type,
            'Net Weight': product.net_weight,
            'Gross Weight': product.gross_weight,

            # Size and Dimensions
            'Size': size,
            'Length': product.length,
            'Width': product.width,
            'Height': product.height,

            # Classification
            'Gender': product.gender,
            'Category': product.category,
            'Usage': product.occasion or 'Casual',
            'Collection': product.collection,
            'Season': product.season or 'All Season',
            'Year': product.year,

            # Business Details
            'HSN': product.hsn_code,
            'Net Quantity': 1,  # Usually 1 for jewelry
            'Inventory': inventory,

            # Marketing
            'Search Tags': search_tags,
            'Occasion': product.occasion,
            'Trend': product.trend,

            # Origin and Manufacturing
            'Country of Origin': product.country_of_origin,
            'Manufacturer': product.manufacturer,
            'Packer': product.packer,
            'Importer': product.importer,

            # Additional
            'Warranty': product.warranty,
            'Certification': product.certification,
            'Care Instructions': product.care_instructions,
        }

        # Initialize image columns as empty (to be filled after upload)
        for i in range(1, 11):
            row_data[f'Image {i}'] = ''

        return row_data

    def generate_excel_file(self, products: List[Product], filename: str = None) -> Path:
        """
        Generate Excel file for Myntra bulk upload.

        Args:
            products: List of Product objects
            filename: Optional custom filename

        Returns:
            Path to generated Excel file
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"myntra_bulk_upload_{timestamp}.xlsx"

        # Ensure .xlsx extension
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'

        output_path = self.output_dir / filename

        # Collect all rows (including variants)
        all_rows = []

        for product in products:
            # If product has variants, create row for each variant
            if product.variants:
                for variant in product.variants:
                    row_data = self.product_to_myntra_row(product, variant)
                    all_rows.append(row_data)
            else:
                # Single product without variants
                row_data = self.product_to_myntra_row(product)
                all_rows.append(row_data)

        # Create DataFrame
        df = pd.DataFrame(all_rows, columns=MYNTRA_TEMPLATE_COLUMNS)

        # Save to Excel with formatting
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Products', index=False)

            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Products']

            # Apply formatting
            self._format_myntra_excel(workbook, worksheet)

        logger.info(f"Generated Myntra Excel file: {output_path}")
        logger.info(f"Total rows: {len(all_rows)}")

        return output_path

    def generate_excel_with_images(self, products: List[Product], filename: str = None) -> Path:
        """
        Generate Excel file and prepare image data for upload.

        Args:
            products: List of Product objects
            filename: Optional custom filename

        Returns:
            Path to generated Excel file
        """
        excel_path = self.generate_excel_file(products, filename)

        # Prepare image mapping for later use
        image_mapping = {}

        for product in products:
            try:
                # Get images for the product
                images = self.image_manager.fetch_images(product.sku)

                if product.variants:
                    # For variants, use the same images
                    for variant in product.variants:
                        image_mapping[variant.sku] = [str(img) for img in images]
                else:
                    image_mapping[product.sku] = [str(img) for img in images]

            except (FileNotFoundError, ValueError) as e:
                logger.warning(f"No images found for SKU {product.sku}: {e}")
                # Continue without images

        # Save image mapping for later use
        mapping_file = excel_path.with_suffix('.json')
        with open(mapping_file, 'w') as f:
            json.dump(image_mapping, f, indent=2)

        logger.info(f"Image mapping saved: {mapping_file}")

        return excel_path

    def schedule_photo_upload_reminder(self, sku: str, delay_minutes: int = 2):
        """
        Schedule a reminder for photo upload after Excel submission.

        Args:
            sku: Product SKU
            delay_minutes: Minutes to wait before reminder
        """
        def reminder_task():
            time.sleep(delay_minutes * 60)  # Convert to seconds
            logger.info(f"🔔 PHOTO UPLOAD REMINDER for SKU: {sku}")
            logger.info("⏰ It's time to upload product photos to Myntra!")

            # Try to get image paths for convenience
            try:
                images = self.image_manager.fetch_images(sku)
                logger.info(f"📸 Ready to upload {len(images)} images:")
                for i, img in enumerate(images[:10], 1):  # Max 10 images
                    logger.info(f"  {i}. {img.name}")
            except Exception as e:
                logger.error(f"Could not fetch images: {e}")

        # Start reminder in background thread
        reminder_thread = threading.Thread(target=reminder_task)
        reminder_thread.daemon = True
        reminder_thread.start()

        self.upload_reminders[sku] = {
            'scheduled_at': datetime.now(),
            'reminder_at': datetime.now() + timedelta(minutes=delay_minutes),
            'thread': reminder_thread
        }

        logger.info(f"Photo upload reminder scheduled for {sku} in {delay_minutes} minutes")

    def get_upload_workflow_instructions(self, excel_path: Path) -> str:
        """
        Get step-by-step instructions for Myntra upload workflow.

        Args:
            excel_path: Path to the generated Excel file

        Returns:
            Formatted instructions string
        """
        instructions = f"""
🔄 MYNTRA BULK UPLOAD WORKFLOW
=============================

Excel File: {excel_path.name}

📋 STEP-BY-STEP INSTRUCTIONS:

1. 📁 LOGIN TO MYNTRA PARTNER PORTAL
   - Go to partner.myntra.com
   - Login with your seller credentials

2. 📊 NAVIGATE TO CATALOG MANAGEMENT
   - Go to Catalog Management > Cataloging
   - Click on "DIY Upload" or "Bulk Upload"

3. 📤 UPLOAD EXCEL FILE
   - Click "Choose File" and select: {excel_path}
   - Click "Upload" and wait for validation
   - Fix any validation errors if shown

4. ⏱️ WAIT FOR PHOTO UPLOAD PROMPT
   - After successful Excel upload, wait exactly 2 minutes
   - A photo upload interface will appear automatically
   - DO NOT refresh the page during this time

5. 📸 UPLOAD PRODUCT PHOTOS
   - The system will show SKUs ready for photo upload
   - Upload images in this order for each product:
     • Image 1: Front view (mandatory)
     • Image 2: Side view
     • Image 3: Back view  
     • Image 4: Close-up/Detail
     • Image 5: Model wearing (if available)
     • Images 6-10: Additional angles

6. ✅ SUBMIT AND REVIEW
   - Click "Submit" after uploading all photos
   - Review the submission status
   - Note any rejected items for correction

⚠️ IMPORTANT NOTES:
- Image size: Max 5MB per image
- Image format: JPG, PNG, WEBP
- Minimum resolution: 1000x1000 pixels
- Maximum 10 images per product
- Don't close browser during upload process

🔔 A reminder will be triggered automatically after 2 minutes!
        """

        return instructions

    def _get_myntra_article_type(self, product_type: ProductType) -> str:
        """Convert ProductType enum to Myntra article type."""
        mapping = {
            ProductType.NECKLACE: 'Necklace and Chains',
            ProductType.BANGLE: 'Bangles and Bracelets',
            ProductType.CHAIN: 'Necklace and Chains',
            ProductType.EARRING: 'Earrings',
            ProductType.RING: 'Rings',
            ProductType.BRACELET: 'Bangles and Bracelets'
        }
        return mapping.get(product_type, 'Jewellery Set')

    def _format_myntra_excel(self, workbook, worksheet):
        """Apply formatting to the Myntra Excel file."""
        from openpyxl.styles import Font, PatternFill, Alignment

        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Apply header formatting
        for col in range(1, len(MYNTRA_TEMPLATE_COLUMNS) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Adjust column widths
        column_widths = {
            'A': 15,  # Style ID
            'B': 40,  # Product Name
            'C': 15,  # Brand Name
            'D': 20,  # Article Type
            'E': 50,  # Product Details
        }

        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width

        # Auto-fit other columns
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            if column_letter not in column_widths:
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 30)  # Max width of 30
                worksheet.column_dimensions[column_letter].width = adjusted_width


# Convenience functions
def generate_myntra_excel(products: List[Product], filename: str = None) -> Path:
    """Generate Myntra Excel file using default exporter."""
    exporter = MyntraExporter()
    return exporter.generate_excel_file(products, filename)


def upload_myntra_bulk(products: List[Product], with_reminders: bool = True) -> Path:
    """
    Complete Myntra upload workflow with reminders.

    Args:
        products: List of products to upload
        with_reminders: Whether to schedule photo upload reminders

    Returns:
        Path to generated Excel file
    """
    exporter = MyntraExporter()

    # Generate Excel with image mapping
    excel_path = exporter.generate_excel_with_images(products)

    # Schedule reminders for each unique SKU
    if with_reminders:
        unique_skus = set()
        for product in products:
            if product.variants:
                unique_skus.update(variant.sku for variant in product.variants)
            else:
                unique_skus.add(product.sku)

        for sku in unique_skus:
            exporter.schedule_photo_upload_reminder(sku)

    # Display workflow instructions
    instructions = exporter.get_upload_workflow_instructions(excel_path)
    print(instructions)

    return excel_path


# For testing
if __name__ == "__main__":
    from data_models import create_sample_necklace, create_sample_bangle

    # Create sample products
    necklace = create_sample_necklace()
    bangle = create_sample_bangle()

    products = [necklace, bangle]

    print("Testing Myntra Excel generation...")

    # Test Excel generation
    exporter = MyntraExporter()
    excel_path = exporter.generate_excel_file(products, "test_myntra_upload.xlsx")

    print(f"✅ Generated Excel file: {excel_path}")

    # Test workflow instructions
    instructions = exporter.get_upload_workflow_instructions(excel_path)
    print(instructions)

    # Test reminder scheduling (with shorter delay for testing)
    print("\nTesting photo upload reminder (5 seconds)...")
    exporter.schedule_photo_upload_reminder("TEST123", delay_minutes=0.083)  # 5 seconds

    # Wait to see the reminder
    time.sleep(6)
    print("Test completed!")
